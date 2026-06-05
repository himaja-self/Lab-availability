"""
Excel Timetable Parser
Handles lab-wise sheets where sheet name = room number (e.g. E-402).
"""

import re
import openpyxl

# ── Constants ────────────────────────────────────────────────
LAB_SHEET_RE = re.compile(r'^[A-Z]-?\s*\d{3}\s*$')

DAY_MAP = {
    'MON': 'MONDAY', 'TUE': 'TUESDAY', 'WED': 'WEDNESDAY',
    'THU': 'THURSDAY', 'FRI': 'FRIDAY', 'SAT': 'SATURDAY',
    'MONDAY': 'MONDAY', 'TUESDAY': 'TUESDAY', 'WEDNESDAY': 'WEDNESDAY',
    'THURSDAY': 'THURSDAY', 'FRIDAY': 'FRIDAY', 'SATURDAY': 'SATURDAY',
}

I_YEAR_SLOTS = [
    ('09:00', '10:00'),
    ('10:00', '11:00'),
    ('11:00', '12:00'),
    ('12:40', '13:40'),
    ('13:40', '14:40'),
    ('14:40', '15:40'),
]

II_III_YEAR_SLOTS = [
    ('10:00', '11:00'),
    ('11:00', '12:00'),
    ('12:00', '13:00'),
    ('13:40', '14:40'),
    ('14:40', '15:40'),
    ('15:40', '16:40'),
]

# Maps Roman numeral prefix to year_group stored in DB
ROMAN_YEAR_MAP = {
    'I': 'I_YEAR',
    'II': 'II_YEAR',
    'III': 'III_YEAR',
    'IV': 'IV_YEAR',
}


# ── Utility Functions ────────────────────────────────────────

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalize_room(room):
    if not room:
        return None
    return re.sub(r'\s+', '', room.upper())


def classify_session(text):
    t = text.upper()
    if 'TRAINING' in t:
        return 'TRAINING'
    if 'MINOR' in t:
        return 'MINOR'
    if 'WORKSHOP' in t:
        return 'WORKSHOP'
    if 'LAB' in t:
        return 'LAB'
    return 'OTHER'


def detect_year_group(text):
    """
    Returns the actual year group: I_YEAR, II_YEAR, III_YEAR, or IV_YEAR.

    I-year:    'I-BRANCH-...' (dash format)
    II-IV year: starts with Roman numeral + space, e.g. 'IV CYS SOC LAB'
    Fallback:  II_YEAR (uses II&III-year time slots)

    The year_group value tells both which year the class belongs to AND
    which time band to use: I_YEAR → I_YEAR_SLOTS, everything else → II_III_YEAR_SLOTS.
    """
    t = text.strip()
    if re.match(r'^I-', t, re.IGNORECASE):
        return 'I_YEAR'
    m = re.match(r'^(II|III|IV|I)\s', t, re.IGNORECASE)
    if m:
        return ROMAN_YEAR_MAP.get(m.group(1).upper(), 'II_YEAR')
    return 'II_YEAR'


def parse_session_text(text):
    """
    Split a raw cell value into (class_name, subject).

    Formats handled:
      I-BRANCH-SECTION-SUBJECT  →  'I BRANCH SECTION', 'SUBJECT'
      I-BRANCH-SUBJECT          →  'I BRANCH', 'SUBJECT'   (e.g. I-AI&DS-PPS LAB)
      ROMAN BRANCH SECTION REST →  'ROMAN BRANCH SECTION', 'REST'
      ROMAN BRANCH REST         →  'ROMAN BRANCH', 'REST'
      fallback                  →  text, text
    """
    t = text.strip()

    # I-year: dash-separated, split on first '-' then handle remainder
    if re.match(r'^I-', t, re.IGNORECASE):
        remainder = t[2:]  # everything after 'I-'
        sub = remainder.split('-')
        branch = sub[0].upper()
        if len(sub) >= 3:
            # I-BRANCH-SECTION-SUBJECT
            section = sub[1].upper()
            subject = '-'.join(sub[2:]).strip().upper()
            return f'I {branch} {section}', subject
        elif len(sub) == 2:
            # I-BRANCH-SUBJECT (no section letter, e.g. I-AI&DS-PPS LAB)
            subject = sub[1].strip().upper()
            return f'I {branch}', subject
        else:
            return f'I {branch}', branch

    # II–IV year with section letter A, B, or C
    m = re.match(r'^(II|III|IV)\s+([\w&]+)\s+([A-C])\s+(.+)', t, re.IGNORECASE)
    if m:
        year = m.group(1).upper()
        branch = m.group(2).upper()
        section = m.group(3).upper()
        subject = m.group(4).strip().upper()
        return f'{year} {branch} {section}', subject

    # II–IV year without section (e.g. 'IV AI&DS DLA LAB', 'III CYS MINOR PIS LAB')
    m2 = re.match(r'^(II|III|IV)\s+([\w&]+)\s+(.+)', t, re.IGNORECASE)
    if m2:
        year = m2.group(1).upper()
        branch = m2.group(2).upper()
        subject = m2.group(3).strip().upper()
        return f'{year} {branch}', subject

    return t, t  # fallback


def get_slot_times(slot_indices, year_group):
    # I_YEAR uses the 9am-start band; all other years use the 10am-start band
    band = I_YEAR_SLOTS if year_group == 'I_YEAR' else II_III_YEAR_SLOTS
    starts = [band[i][0] for i in slot_indices if i < len(band)]
    ends = [band[i][1] for i in slot_indices if i < len(band)]
    if not starts:
        return None, None
    return starts[0], ends[-1]


def get_merged_span(ws, row, col):
    for merge in ws.merged_cells.ranges:
        if merge.min_row <= row <= merge.max_row and \
           merge.min_col <= col <= merge.max_col:
            return merge.min_col, merge.max_col
    return col, col


# ── Lab-Wise Sheet Parser ────────────────────────────────────

def parse_lab_wise_sheet(ws, room_number, source_file):
    sessions = []

    # Find header row (contains slot labels I, II, III)
    header_row = None
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_vals = [clean(ws.cell(row_idx, c).value) for c in range(1, 15)]
        non_none = [v for v in row_vals if v]
        if 'I' in non_none and 'II' in non_none and 'III' in non_none:
            header_row = row_idx
            break

    if not header_row:
        return sessions

    # Identify slot columns
    slot_order = []
    for col in range(1, ws.max_column + 1):
        val = clean(ws.cell(header_row, col).value)
        if val in ('I', 'II', 'III', 'IV', 'V', 'VI'):
            slot_order.append((col, val))
    slot_order.sort(key=lambda x: x[0])

    if not slot_order:
        return sessions

    slot_label_to_idx = {label: idx for idx, (col, label) in enumerate(slot_order)}

    # Find day rows (skip time-band rows right after header)
    day_rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell_val = clean(ws.cell(row_idx, 1).value) or clean(ws.cell(row_idx, 2).value)
        if cell_val:
            day_key = cell_val.upper()[:3]
            if day_key in DAY_MAP:
                day_rows.append((row_idx, DAY_MAP[day_key]))

    # Parse each day row
    for row_idx, day_name in day_rows:
        processed_cols = set()

        for col, slot_label in slot_order:
            if col in processed_cols:
                continue

            val = clean(ws.cell(row_idx, col).value)
            if not val:
                continue

            # Skip lunch or time-string cells that leaked into day rows
            if 'LUNCH' in val.upper() or re.search(r'\d{1,2}:\d{2}\s*(AM|PM)', val, re.IGNORECASE):
                continue

            min_col, max_col = get_merged_span(ws, row_idx, col)

            covered_slot_indices = [
                slot_label_to_idx[lbl]
                for c, lbl in slot_order
                if min_col <= c <= max_col
            ]

            for c, lbl in slot_order:
                if min_col <= c <= max_col:
                    processed_cols.add(c)

            year_group = detect_year_group(val)
            start_time, end_time = get_slot_times(covered_slot_indices, year_group)

            if not start_time:
                continue

            class_name, subject = parse_session_text(val)

            sessions.append({
                'room_number': normalize_room(room_number),
                'day_of_week': day_name,
                'start_time': start_time,
                'end_time': end_time,
                'class_name': class_name,
                'subject': subject,
                'session_type': classify_session(val),
                'year_group': year_group,
                'source_file': source_file,
            })

    return sessions


def parse_lab_wise_excel(wb, source_file):
    all_sessions = []
    parsed_sheets = []
    skipped_sheets = []

    for sheet_name in wb.sheetnames:
        if LAB_SHEET_RE.match(sheet_name.strip()):
            ws = wb[sheet_name]
            room = normalize_room(sheet_name.strip())
            sessions = parse_lab_wise_sheet(ws, room, source_file)
            all_sessions.extend(sessions)
            parsed_sheets.append(sheet_name)
        else:
            skipped_sheets.append(sheet_name)

    return all_sessions, parsed_sheets, skipped_sheets


# ── Main Entry Point ─────────────────────────────────────────

def parse_excel(filepath, source_file):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    result = {
        'sessions': [],
        'labs_found': [],
        'parsed_sheets': [],
        'skipped_sheets': [],
        'warnings': [],
        'content_types_found': [],
    }

    lab_sessions, parsed, skipped = parse_lab_wise_excel(wb, source_file)
    if lab_sessions:
        result['sessions'].extend(lab_sessions)
        result['parsed_sheets'].extend(parsed)
        result['content_types_found'].append('lab_wise')
        result['labs_found'] = sorted({s['room_number'] for s in lab_sessions})

    result['skipped_sheets'].extend(skipped)
    return result


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else 'Lab_timetables.xlsx'
    result = parse_excel(path, path.split('/')[-1])
    print(f"Sessions:      {len(result['sessions'])}")
    print(f"Labs found:    {result['labs_found']}")
    print(f"Parsed sheets: {result['parsed_sheets']}")
    print()
    for s in result['sessions'][:20]:
        print(f"  {s['room_number']:<8} {s['day_of_week']:<10} {s['start_time']}-{s['end_time']}  "
              f"{s['class_name']:<20} | {s['subject']:<20} [{s['year_group']}]")
    if len(result['sessions']) > 20:
        print(f"  ... and {len(result['sessions'])-20} more")